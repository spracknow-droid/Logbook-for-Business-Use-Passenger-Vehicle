import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import holidays
from datetime import datetime, timedelta
import requests
# import copy # copy 모듈은 더 이상 필요 없으므로 주석 처리하거나 삭제

# ----------------- GitHub에서 템플릿 파일 로드 -----------------
github_template_url = "https://raw.githubusercontent.com/spracknow-droid/Logbook-for-Business-Use-Passenger-Vehicle/main/Logbook-for-Business-Use-Passenger-Vehicle-Template(2025).xlsx"

@st.cache_data
def load_template_from_github(url):
    """
    GitHub의 Raw URL에서 엑셀 파일을 다운로드하여 openpyxl 워크북 객체와 그 내용을 담은 BytesIO 객체를 반환합니다.
    """
    try:
        response = requests.get(url)
        if response.status_code == 200:
            # BytesIO 객체에 원본 파일 내용 저장
            template_bytes = BytesIO(response.content)
            # BytesIO 객체에서 워크북 객체 로드
            return load_workbook(template_bytes), template_bytes
        else:
            st.error(f"GitHub에서 파일을 불러오는 데 실패했습니다. 상태 코드: {response.status_code}")
            return None, None
    except requests.exceptions.RequestException as e:
        st.error(f"네트워크 오류가 발생했습니다: {e}")
        return None, None

# ----------------- 기존 로직 -----------------
# 국가별 공휴일 설정 (예: 한국)
kr_holidays = holidays.KR()

# 한국어 요일 맵핑
KOREAN_WEEKDAYS = {
    0: '월',
    1: '화',
    2: '수',
    3: '목',
    4: '금',
    5: '토',
    6: '일'
}

def is_working_day(date):
    """주어진 날짜가 평일(월~금)이고 공휴일이 아닌지 확인"""
    if date.weekday() >= 5 or date in kr_holidays:
        return False
    return True

def get_next_working_day(date):
    """가장 가까운 평일(working day)을 찾기"""
    while not is_working_day(date):
        date += timedelta(days=1)
    return date

def get_total_working_days(start_date, end_date):
    """시작일자와 종료일자 사이의 총 평일(working days) 수 계산"""
    total_days = 0
    current_date = start_date
    while current_date <= end_date:
        if is_working_day(current_date):
            total_days += 1
        current_date += timedelta(days=1)
    return total_days

def fill_worksheet_data(ws, data):
    """시트에 데이터를 채워넣는 공통 함수"""
    # B, C, D열의 너비를 고정된 값으로 설정
    ws.column_dimensions['B'].width = 18.00
    ws.column_dimensions['C'].width = 8.00
    ws.column_dimensions['D'].width = 8.00

    # 엑셀 셀에 데이터 입력
    ws['B9'] = data['차종']
    ws['E9'] = data['자동차등록번호']
    ws['C15'] = data['부서']
    ws['E15'] = data['성명']
    
    # 숫자형 데이터를 정수로 변환 후 포맷 적용
    ws['G15'] = int(data['시작주행거리'])
    ws['G15'].number_format = '#,##0'
    ws['K15'] = int(data['일평균주행거리'])
    ws['K15'].number_format = '#,##0'
    
    # I15 셀 계산: G15 + K15
    ws['I15'] = int(data['시작주행거리'] + data['일평균주행거리'])
    ws['I15'].number_format = '#,##0'

    # 운행기록부 데이터 입력
    b_col_idx = 2  # 'B'
    
    start_row = 15
    end_row = 264

    # 이전 주행 후 계기판 거리 초기값 설정
    prev_driving_distance = data['시작주행거리']

    # B열(날짜) 채우기 및 운행기록부 데이터 입력
    current_date = data['사용시작일자']
    row_num = start_row
    while row_num <= end_row and current_date <= data['사용종료일자']:
        if is_working_day(current_date):
            # 1. 사용일자(요일)를 한국어로 변경하고 왼쪽 정렬
            weekday_korean = KOREAN_WEEKDAYS.get(current_date.weekday(), '')
            cell_date_value = current_date.strftime(f'%Y-%m-%d({weekday_korean})')
            date_cell = ws.cell(row=row_num, column=b_col_idx, value=cell_date_value)
            date_cell.alignment = Alignment(horizontal='left')
            
            # 2. 부서와 성명 값 채우기
            ws.cell(row=row_num, column=3, value=data['부서'])
            ws.cell(row=row_num, column=5, value=data['성명'])
            ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='left')
            ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='left')

            # 3. 주행전 계기판 거리를 이전 값으로 설정하고 정수 및 콤마 포맷 적용
            ws.cell(row=row_num, column=7, value=int(prev_driving_distance)).number_format = '#,##0'
            
            # 4. 주행 후 계기판 거리 계산
            driving_distance = prev_driving_distance + data['일평균주행거리']
            
            # 5. 최종 주행거리를 초과하지 않도록 설정
            if driving_distance > data['최종주행거리']:
                driving_distance = data['최종주행거리']
            
            # 6. 주행 후 계기판 거리에 값 입력 및 정수, 콤마 포맷 적용
            ws.cell(row=row_num, column=9, value=int(driving_distance)).number_format = '#,##0'
            
            # 7. 일일 주행거리(K열) 계산 및 값 입력
            daily_mileage = ws.cell(row=row_num, column=9).value - ws.cell(row=row_num, column=7).value
            ws.cell(row=row_num, column=11, value=int(daily_mileage)).number_format = '#,##0'

            # 8. 출퇴근용 업무 사용거리(L열)에 일일 주행거리(K열)와 동일한 값 입력
            ws.cell(row=row_num, column=12, value=int(daily_mileage)).number_format = '#,##0'

            # 9. 다음 행의 '주행전 계기판 거리'를 위해 현재 행의 '주행 후 계기판 거리'를 저장
            prev_driving_distance = driving_distance

            row_num += 1
        current_date += timedelta(days=1)

    # 1. G266셀에 K15:K264의 합계 입력
    ws['G266'] = f'=SUM(K15:K264)'
    ws['G266'].number_format = '#,##0'

    # 2. L266셀에 L15:L264의 합계 입력
    ws['L266'] = f'=SUM(L15:L264)'
    ws['L266'].number_format = '#,##0'

    # 3. Q266셀에 L266/G266의 백분율 입력 및 포맷 설정
    ws['Q266'] = f'=L266/G266'
    ws['Q266'].number_format = '#,###%'
    
    return ws

# ----------------- Streamlit 앱 시작 -----------------
st.set_page_config(layout="wide")
st.title("🚙 업무용 승용차 운행기록부 생성기")

# 템플릿 파일과 BytesIO 객체 동시 로드
template_workbook, template_bytes_io = load_template_from_github(github_template_url)

if template_workbook is None:
    st.warning("GitHub에서 템플릿 파일을 불러오는 데 실패했습니다. URL을 확인해 주세요.")
else:
    st.info("GitHub에서 템플릿 파일을 성공적으로 불러왔습니다.")

    # 탭 생성
    tab1, tab2 = st.tabs(["운행기록부 단일 생성", "운행기록부 다중 생성"])

    # 탭 1의 내용
    with tab1:
        # 현재 연도
        current_year = datetime.now().year
        default_start_date = datetime(current_year, 1, 1).date()
        default_end_date = datetime(current_year, 12, 31).date()

        st.header("1. 차량 정보 입력")
        col1, col2 = st.columns(2)
        with col1:
            car_model = st.text_input("차종", "쏘나타")
        with col2:
            car_number = st.text_input("자동차등록번호", "12가3456")

        st.header("2. 기간 및 운전자 정보")
        col3, col4, col5 = st.columns(3)
        with col3:
            start_date = st.date_input("사용 시작일자", default_start_date)
        with col4:
            end_date = st.date_input("사용 종료일자", default_end_date)
        with col5:
            department = st.text_input("부서", "총무부")
            name = st.text_input("성명", "김철수")

        st.header("3. 주행 정보")
        col6, col7 = st.columns(2)
        with col6:
            start_mileage = st.number_input("시작 주행거리 (km)", value=0.0, step=0.1)
        with col7:
            # end_mileage의 기본값이 start_mileage보다 크거나 같도록 동적으로 설정합니다.
            # 사용자가 start_mileage를 15000.0보다 높게 입력해도 문제가 없습니다.
            end_mileage_default = max(start_mileage, 15000.0)
            end_mileage = st.number_input(
                "최종 주행거리 (km)",
                value=end_mileage_default,
                step=0.1,
                min_value=start_mileage
            )

        if st.button("운행기록부 생성"):
            total_mileage = end_mileage - start_mileage
            total_working_days = get_total_working_days(start_date, end_date)
            
            if total_working_days == 0:
                st.warning("선택한 기간에 평일(Working Day)이 없습니다. 날짜를 다시 선택해주세요.")
            else:
                avg_daily_mileage = total_mileage / total_working_days

                start_date_obj = datetime.combine(start_date, datetime.min.time())
                end_date_obj = datetime.combine(end_date, datetime.min.time())

                report_data = {
                    "차종": car_model,
                    "자동차등록번호": car_number,
                    "부서": department,
                    "성명": name,
                    "사용시작일자": get_next_working_day(start_date_obj),
                    "사용종료일자": end_date_obj,
                    "시작주행거리": start_mileage,
                    "최종주행거리": end_mileage,
                    "총주행거리": total_mileage,
                    "총근무일수": total_working_days,
                    "일평균주행거리": avg_daily_mileage
                }
                
                # BytesIO 객체에서 새로운 워크북 로드 (가장 안전한 방법)
                template_bytes_io.seek(0)
                workbook = load_workbook(template_bytes_io)
                
                ws = workbook.active
                ws.title = f"{report_data['성명']}_{report_data['자동차등록번호']}"
                fill_worksheet_data(ws, report_data)
                
                excel_buffer = BytesIO()
                workbook.save(excel_buffer)
                excel_buffer.seek(0)
                
                st.success("운행기록부 작성이 완료되었습니다!")
                
                st.download_button(
                    label="다운로드 (엑셀 파일)",
                    data=excel_buffer,
                    file_name="업무용_승용차_운행기록부.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    # 탭 2의 내용
    with tab2:
        st.header("운행기록부 다중 생성")
        st.info("다중 생성을 위해 아래 예시와 동일한 형식의 엑셀 파일을 업로드하세요.")
        
        example_data = {
            '차종': ['쏘나타', '그랜저', '아반떼'],
            '자동차등록번호': ['12가3456', '78나9012', '34다5678'],
            '사용 시작일자': ['2023-01-01', '2023-03-01', '2023-05-01'],
            '사용 종료일자': ['2023-06-30', '2023-08-31', '2023-10-31'],
            '부서': ['총무부', '영업부', '개발부'],
            '성명': ['김철수', '이영희', '박지훈'],
            '시작 주행거리': [10000, 15000, 20000],
            '최종 주행거리': [12500, 18000, 22500]
        }
        example_df = pd.DataFrame(example_data)
        st.dataframe(example_df, hide_index=True)
        
        uploaded_file_multi = st.file_uploader("다중 생성용 엑셀 파일을 업로드하세요", type=["xlsx"], key="multi_upload")

        if uploaded_file_multi:
            st.success("다중 생성용 엑셀 파일이 성공적으로 업로드되었습니다.")
            
            if st.button("운행기록부 일괄 생성"):
                df = pd.read_excel(uploaded_file_multi)
                
                progress_bar = st.progress(0)
                total_rows = len(df)
                
                # BytesIO 객체에서 새로운 워크북 로드
                # 원본 BytesIO 객체의 포인터를 처음으로 되돌려야 다시 읽을 수 있음
                template_bytes_io.seek(0)
                new_workbook = load_workbook(template_bytes_io)

                # 첫 번째 기본 시트의 이름을 임시로 저장
                default_sheet_name = new_workbook.active.title

                for i in range(total_rows):
                    row = df.iloc[i]
                    
                    total_mileage = row['최종 주행거리'] - row['시작 주행거리']
                    start_date_obj = row['사용 시작일자'].to_pydatetime()
                    end_date_obj = row['사용 종료일자'].to_pydatetime()
                    total_working_days = get_total_working_days(start_date_obj, end_date_obj)
                    
                    if total_working_days == 0:
                        st.warning(f"데이터 파일의 {i+1}번째 행에 해당하는 기간에 평일이 없습니다. 해당 시트는 생성되지 않습니다.")
                        progress_bar.progress((i + 1) / total_rows)
                        continue
                        
                    avg_daily_mileage = total_mileage / total_working_days
                    
                    data = {
                        "차종": row['차종'],
                        "자동차등록번호": row['자동차등록번호'],
                        "부서": row['부서'],
                        "성명": row['성명'],
                        "사용시작일자": get_next_working_day(start_date_obj),
                        "사용종료일자": end_date_obj,
                        "시작주행거리": row['시작 주행거리'],
                        "최종주행거리": row['최종 주행거리'],
                        "총주행거리": total_mileage,
                        "총근무일수": total_working_days,
                        "일평균주행거리": avg_daily_mileage
                    }
                    
                    # 새 시트 복사 및 데이터 입력
                    ws = new_workbook.copy_worksheet(new_workbook.active)
                    ws.title = f"{data['성명']}_{data['자동차등록번호']}"
                    fill_worksheet_data(ws, data)
                    
                    progress_bar.progress((i + 1) / total_rows)
                
                # 첫 번째 기본 시트 제거
                new_workbook.remove(new_workbook[default_sheet_name])
                
                excel_buffer = BytesIO()
                new_workbook.save(excel_buffer)
                excel_buffer.seek(0)

                st.success("운행기록부 일괄 작성이 완료되었습니다!")
                
                st.download_button(
                    label="다운로드 (엑셀 파일)",
                    data=excel_buffer,
                    file_name="업무용_승용차_운행기록부_일괄생성.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
